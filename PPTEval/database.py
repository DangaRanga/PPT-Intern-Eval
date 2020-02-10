import xlsxwriter
import openpyxl
import smtplib
import pandas as pd
import os
from validatorfunctions import *
import smtplib
from email.mime.text import MIMEText
import pyowm

#------------------------------------------------------------------------------#
# Function to make the excel workbook as an excel macro, since openpyxl
# corrupts xlsx files
#-------------------------------------------------------------------------------#
def make_xlsm_wb():
    file = openpyxl.load_workbook('WorkerDB.xlsx',read_only=False,keep_vba=True)
    file.save('WorkerDB.xlsm')

# ---------------------------------------------------------------------------- #
# Function to make the layout for the database using excel
# ---------------------------------------------------------------------------- #
def make_database():
    df = pd.DataFrame({'Name':[] ,
                        'Address1':[],
                        'City':[],
                        'Country':[],
                        'Telephone':[],
                        'Role':[],
                        'Email':[],
                        'AddressLocation':[],
                        'CountryCode':[]})
    df.to_excel("WorkerDB.xlsx")


#----------------------------------------------------------------------------- #
# Function to add the name to the spreadsheet
# ---------------------------------------------------------------------------- #

def add_name(sheet_name):
    Terminate = False
    na_row = 2
    na_col = 2
    while True:
        name = input("Enter the individual's name:  ")
        if is_empty(name):
            print("Enter a valid name")
            continue
        else:
            break
    if sheet_name.cell(row=na_row,column=na_col).value != None:
        while sheet_name.cell(row=na_row,column=na_col).value != None:
                na_row += 1
        sheet_name.cell(row=na_row,column=na_col).value = name
    else:
        sheet_name.cell(row=na_row,column=na_col).value = name

#----------------------------------------------------------------------------- #
# Function to add the address to the spreadsheet
#----------------------------------------------------------------------------- #
def add_address(sheet_name):
    add_row = 2
    add_column = 3
    while True:
        address = input("Enter the individual's address: ")
        if is_empty(address):
            print("Please enter a valid address")
            continue
        else:
            break
    if sheet_name.cell(row=add_row,column=add_column).value != None:
        while sheet_name.cell(row=add_row,column=add_column).value != None:
                add_row += 1
        sheet_name.cell(row=add_row,column=add_column).value = address
    else:
        sheet_name.cell(row=add_row,column=add_column).value = address

#----------------------------------------------------------------------------- #
# Function to add the city to the spreadsheet
#----------------------------------------------------------------------------- #
def add_city(sheet_name):
    city_row = 2
    city_column = 4
    while True:
        city = input("Enter the individual's city: ")
        if is_empty(city):
            print("Please enter a valid city")
            continue
        else:
            break
    if sheet_name.cell(row=city_row,column=city_column).value != None:
        while sheet_name.cell(row=city_row,column=city_column).value != None:
                city_row += 1
        sheet_name.cell(row=city_row,column=city_column).value = city
    else:
        sheet_name.cell(row=city_row,column=city_column).value = city

#----------------------------------------------------------------------------- #
# Function to add the country of the individual to the spreadsheet
#----------------------------------------------------------------------------- #
def add_country(sheet_name):
    country_row = 2
    country_column = 5
    while True:
        country = input("Enter the individual's country: ")
        if is_empty(country):
            print("Please enter a valid country")
            continue
        else:
            break
    if sheet_name.cell(row=country_row,column=country_column).value != None:
        while sheet_name.cell(row=country_row,column=country_column).value != None:
                country_row += 1
        sheet_name.cell(row=country_row,column=country_column).value = country
    else:
        sheet_name.cell(row=country_row,column=country_column).value = country

#----------------------------------------------------------------------------- #
# Function to add the telephone number of the individual to the spreadsheet
#----------------------------------------------------------------------------- #
def add_telephone_no(sheet_name):
    tele_no_row = 2
    tele_no_column = 6
    while True:
        tele_no = input("Enter the individual's telephone number: ")
        if is_empty(tele_no):
            print("Please enter a valid telephone number")
            continue
        else:
            break
    if sheet_name.cell(row=tele_no_row,column=tele_no_column).value != None:
        while sheet_name.cell(row=tele_no_row,column=tele_no_column).value != None:
                tele_no_row += 1
        sheet_name.cell(row=tele_no_row,column=tele_no_column).value = tele_no
    else:
        sheet_name.cell(row=tele_no_row,column=tele_no_column).value = tele_no
#----------------------------------------------------------------------------- #
# Function to add the city to the spreadsheet
#----------------------------------------------------------------------------- #
def add_role(sheet_name):
    role_row = 2
    role_column = 7
    while True:
        role = input("Enter the individual's role: ")
        if is_empty(role):
            print("Please enter a valid role")
            continue
        else:
            break
    if sheet_name.cell(row=role_row,column=role_column).value != None:
        while sheet_name.cell(row=role_row,column=role_column).value != None:
                role_row += 1
        sheet_name.cell(row=role_row,column=role_column).value = role
    else:
        sheet_name.cell(row=role_row,column=role_column).value = role

# -----------------------------------------------------------------------------#
# Function to add the email entered to the spreadsheet
# -----------------------------------------------------------------------------#

def add_email(sheet_name):
    Terminate = False
    em_row = 2
    em_col = 8
    while True:
        email = input("Enter the individual's email address: ")
        if is_empty(email):
            print("Please enter a valid email address")
            continue
        else:
            break
    if sheet_name.cell(row=em_row,column=em_col).value != None:
        while sheet_name.cell(row=em_row,column=em_col).value != None:
                em_row += 1
        sheet_name.cell(row=em_row,column=em_col).value = email
    else:
        sheet_name.cell(row=em_row,column=em_col).value = email

#------------------------------------------------------------------------------#
# Function to add the address location to the spreadsheet
# -----------------------------------------------------------------------------#

def add_address_location(sheet_name):
    add_loc_row = 2
    add_loc_column = 9
    while True:
        add_loc = input("Enter the individual's address location: ")
        if is_empty(add_loc):
            print("Please enter a valid address location")
            continue
        else:
            break
    if sheet_name.cell(row=add_loc_row,column=add_loc_column).value != None:
        while sheet_name.cell(row=add_loc_row,column=add_loc_column).value != None:
                add_loc_row += 1
        sheet_name.cell(row=add_loc_row,column=add_loc_column).value = add_loc
    else:
        sheet_name.cell(row=add_loc_row,column=add_loc_column).value = add_loc



#------------------------------------------------------------------------------#
# Functions to retrieve data from the spreadsheet and store them as iterables
# With the use of list comprehensions
#------------------------------------------------------------------------------ #

#------------------------------------------------------------------------------#
# Function to get the names for each worker
# -----------------------------------------------------------------------------#

def names(sheet_name):
    name_lst =  [name.value for name in sheet_name['B']]
    return name_lst[1:] # This is done to prevent the header from being stored

#------------------------------------------------------------------------------#
# Function to get the worker roles
#------------------------------------------------------------------------------#
def roles(sheet_name):
    role_lst = [role.value for role in sheet_name['G'] if role.value == 'IT' or role.value == 'IT Worker']
    return role_lst

#------------------------------------------------------------------------------#
# Function to get the locations by city for each worker
# -----------------------------------------------------------------------------#

def cities(sheet_name):
    city_lst = [city.value for city in sheet_name['D']]
    return city_lst[1:] # This is done to prevent the header from being stored

#------------------------------------------------------------------------------#
# Function to get the email addresses of each worker
# -----------------------------------------------------------------------------#

def emails(sheet_name):
    email_lst = []
    for email in sheet_name['H']:
        email_lst.append(email.value)
    return email_lst[1:]

# -----------------------------------------------------------------------------#
# Function to store the cities and emails together as an iterable object
# -----------------------------------------------------------------------------#

def city_mail(sheet_name):
    return list(zip(cities(sheet_name),emails(sheet_name)))

#----------------------------------------------------------------------------- #
# Function to make a dictionary to store the names and a dictionary of the
# cities and emails
# ---------------------------------------------------------------------------- #
def name_dict(sheet_name):
    name_lst = names(sheet_name)
    city_mails = city_mail(sheet_name)
    return dict(zip(name_lst,city_mails))

def email_location_dict(sheet_name):
    return dict(city_mail(sheet_name))

def role_email_dict(sheet_name):
    worker_roles = roles(sheet_name)
    email_ad = emails(sheet_name)
    return dict(zip(worker_roles,email_ad))
#----------------------------------------------------------------------------- #
# Function to make entries in the spreadsheet using the application
# ---------------------------------------------------------------------------- #
def make_entry(sheet_name):
    Terminate = False
    file = openpyxl.load_workbook('WorkerDB.xlsm',read_only=False,keep_vba=True)
    while not Terminate:
         print("# --------------------Database Entry Screen---------------------- #")
         add_name(sheet_name)
         add_address(sheet_name)
         add_city(sheet_name)
         add_country(sheet_name)
         add_telephone_no(sheet_name)
         add_role(sheet_name)
         add_email(sheet_name)
         add_address_location(sheet_name)
         choice = input("Would you like to continue adding entries to the database? (Y/N) ")
         if choice == "Y":
             continue
         else:
            Terminate = True
    print("Saving updates to database")
#------------------------------------------------------------------------------#
# Getter/Accessor functions for the database
#------------------------------------------------------------------------------#
def get_city(name,sheet_name):
    return name_dict(sheet_name)[name][0]
def get_email(name,sheet_name):
    return name_dict(sheet_name)[name][1]
#------------------------------------------------------------------------------#
# Function to create the mailing list based on the location
#------------------------------------------------------------------------------#
def mailing_list_loc(sheet_name,location):
    email_lst = city_mail(sheet_name)
    mailing_list = [city[1] for city in email_lst if city[0] == location]
    return mailing_list
#-----------------------------------------------------------------------------#
